"""Regression tests for Foreign Assets report summaries."""

import csv
from datetime import date

from openpyxl import load_workbook

from equitywise.calculators.fa_calculator import FADeclarationSummary
from equitywise.calculators.rsu_calculator import (
    RSUCalculationSummary,
    SaleEvent,
)
from equitywise.reports.csv_reporter import CSVReporter
from equitywise.reports.excel_reporter import ExcelReporter


def test_fa_reports_use_peak_balance_for_declaration_requirement(tmp_path):
    summary = FADeclarationSummary(
        declaration_date=date(2026, 7, 18),
        calendar_year="2025",
        closing_balance_inr=150_000.0,
        peak_balance_inr=300_000.0,
        vested_holdings_inr=150_000.0,
    )

    csv_file = CSVReporter(tmp_path).generate_fa_report(
        summary, calendar_year="2025", detailed=False
    )[0]
    with csv_file.open(newline="", encoding="utf-8") as handle:
        declaration_row = next(
            row for row in csv.reader(handle) if row[1] == "Declaration Required?"
        )
    assert declaration_row[2] == "YES"

    excel_file = ExcelReporter(tmp_path).generate_fa_report(
        summary, calendar_year="2025", detailed=False
    )
    workbook = load_workbook(excel_file, data_only=True)
    rows = workbook["FA Summary"].iter_rows(values_only=True)
    declaration_row = next(row for row in rows if row[1] == "Declaration Required?")
    assert declaration_row[2] == "YES"


def test_rsu_excel_sale_totals_align_with_two_exchange_rate_columns(tmp_path):
    sale = SaleEvent(
        sale_date=date(2025, 5, 2),
        acquisition_date=date(2025, 4, 15),
        grant_date=date(2025, 4, 15),
        grant_number="RU123",
        order_number="ORDER-1",
        quantity_sold=3,
        sale_price_usd=380.45,
        sale_proceeds_usd=1141.35,
        sale_proceeds_inr=95709.96,
        cost_basis_usd=1052.15,
        cost_basis_inr=90243.54,
        capital_gain_usd=79.82,
        capital_gain_inr=4679.85,
        gain_type="Short-term",
        exchange_rate_sale=84.25,
        acquisition_exchange_rate=85.10,
        cost_basis_exchange_rate=85.10,
        vest_exchange_rate=85.7706,
        financial_year="FY25-26",
        gross_capital_gain_usd=89.20,
        gross_capital_gain_inr=5466.43,
        sale_expense_usd=9.38,
        sale_expense_inr=786.58,
    )
    summary = RSUCalculationSummary(financial_year="FY25-26")

    excel_file = ExcelReporter(tmp_path).generate_rsu_report(
        summary,
        vesting_events=[],
        sale_events=[sale],
        financial_year="FY25-26",
    )
    worksheet = load_workbook(excel_file, data_only=True)["Sale Events"]

    headers = [cell.value for cell in worksheet[3]]
    total_values = [cell.value for cell in worksheet[5]]
    assert headers[5:9] == [
        "Sale Rule 115 SBI TTBR",
        "Cost Basis Conversion Rate",
        "Sale Proceeds (USD)",
        "Sale Proceeds (INR)",
    ]
    assert total_values[5:13] == [
        "-",
        "-",
        sale.sale_proceeds_usd,
        sale.sale_proceeds_inr,
        sale.cost_basis_usd,
        sale.cost_basis_inr,
        sale.capital_gain_usd,
        sale.capital_gain_inr,
    ]
    assert headers[20:22] == [
        "Acquisition SBI TTBR (Prior Month)",
        "Capital Gains Calculation Method",
    ]
    assert total_values[20:22] == ["-", "-"]
