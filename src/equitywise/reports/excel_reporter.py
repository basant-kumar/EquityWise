"""Excel report generation for RSU and FA calculations."""

import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from loguru import logger

from ..calculators.rsu_calculator import RSUCalculationSummary, VestingEvent, SaleEvent
from ..calculators.fa_calculator import FADeclarationSummary, EquityHolding, VestWiseDetails


class ExcelReporter:
    """Excel report generator for RSU and FA calculations."""
    
    def __init__(self, output_dir: str = "output"):
        """Initialize Excel reporter.
        
        Args:
            output_dir: Directory to save Excel reports
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # Excel styling
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.currency_format = '₹#,##0.00'
        self.number_format = '#,##0.00'
        self.date_format = 'DD/MM/YYYY'
        
        # Borders
        thin_border = Side(border_style="thin", color="000000")
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    def generate_rsu_report(
        self,
        summary: RSUCalculationSummary,
        vesting_events: List[VestingEvent],
        sale_events: List[SaleEvent],
        bank_transactions: Optional[List[Dict]] = None,
        financial_year: str = None,
        detailed: bool = True
    ) -> Path:
        """Generate comprehensive RSU Excel report.
        
        Args:
            summary: RSU calculation summary
            vesting_events: List of vesting events
            sale_events: List of sale events  
            bank_transactions: Bank statement transactions
            financial_year: Financial year for the report
            detailed: Include detailed breakdowns
            
        Returns:
            Path to generated Excel file
        """
        logger.info(f"Generating RSU Excel report for {financial_year or 'all years'}")
        
        # Create filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fy_suffix = f"_{financial_year}" if financial_year else ""
        filename = f"RSU_Report{fy_suffix}_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_rsu_summary_sheet(wb, summary, financial_year)
        
        if detailed:
            if vesting_events:
                self._create_vesting_events_sheet(wb, vesting_events)
            if sale_events:
                self._create_sale_events_sheet(wb, sale_events)
            if bank_transactions:
                self._create_bank_reconciliation_sheet(wb, sale_events, bank_transactions)
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"RSU Excel report saved: {filepath}")
        
        return filepath
    
    def generate_fa_report(
        self,
        summary: FADeclarationSummary,
        equity_holdings: List[EquityHolding] = None,
        vest_wise_details: List[VestWiseDetails] = None,
        calendar_year: str = None,
        detailed: bool = True
    ) -> Path:
        """Generate comprehensive FA Excel report.
        
        Args:
            summary: FA declaration summary
            equity_holdings: Current equity holdings
            vest_wise_details: Vest-wise breakdown details
            calendar_year: Calendar year for the report
            detailed: Include detailed breakdowns
            
        Returns:
            Path to generated Excel file
        """
        logger.info(f"Generating FA Excel report for {calendar_year or 'all years'}")
        
        # Create filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        cy_suffix = f"_{calendar_year}" if calendar_year else ""
        filename = f"FA_Report{cy_suffix}_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_fa_summary_sheet(wb, summary, calendar_year)
        self._create_company_details_sheet(wb)  # Add company & account details sheet
        
        if detailed:
            if equity_holdings:
                self._create_equity_holdings_sheet(wb, equity_holdings)
            if vest_wise_details:
                self._create_vest_wise_details_sheet(wb, vest_wise_details)
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"FA Excel report saved: {filepath}")
        
        return filepath
    
    def _create_rsu_summary_sheet(self, wb: Workbook, summary: RSUCalculationSummary, financial_year: str):
        """Create RSU summary sheet."""
        ws = wb.create_sheet("RSU Summary")
        
        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f"RSU Tax Summary - {financial_year or 'All Years'}"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Vesting Details Section
        vesting_data = [
            ["", "Vesting Details", "USD Amount", "INR Amount"],
            ["", "Total Vested Shares", f"{summary.total_vested_quantity:.0f} shares", ""],
            ["", "Total Vesting Income", f"${summary.total_taxable_gain_usd:,.2f}", f"₹{summary.total_taxable_gain_inr:,.2f}"],
            ["", "Average Exchange Rate", f"₹{summary.average_exchange_rate:.4f}/USD", ""],
            ["", "", "", ""],
        ]
        
        # Sale Details Section
        sale_data = [
            ["", "Sale Details", "USD Amount", "INR Amount"],
            ["", "Total Sold Shares", f"{summary.total_sold_quantity:.0f} shares", ""],
            ["", "Total Purchase Amount", f"${summary.total_cost_basis_usd:,.2f}", f"₹{summary.total_cost_basis_inr:,.2f}"],
            ["", "Gross Sale Proceeds", f"${summary.total_sale_proceeds_usd:,.2f}", f"₹{summary.total_sale_proceeds_inr:,.2f}"],
            ["", "Deductible Sale Expenses", f"${summary.total_sale_expenses_usd:,.2f}", f"₹{summary.total_sale_expenses_inr:,.2f}"],
            ["", "", "", ""],
            ["", "Total Capital Gains", f"${summary.total_capital_gains_usd:,.2f}", f"₹{summary.total_capital_gains_inr:,.2f}"],
            ["", "Short-term Gains", f"${summary.short_term_gains_usd:,.2f}", f"₹{summary.short_term_gains_inr:,.2f}"],
            ["", "Long-term Gains", f"${summary.long_term_gains_usd:,.2f}", f"₹{summary.long_term_gains_inr:,.2f}"],
        ]
        
        # Combine all sections
        summary_data = vesting_data + sale_data
        
        # Add data to sheet
        for row_idx, row_data in enumerate(summary_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Style section headers (Vesting Details, Sale Details)
                if "Vesting Details" in str(value) or "Sale Details" in str(value):
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                
                cell.border = self.border
        
        # Auto-adjust column widths with improved calculation
        self._auto_adjust_column_widths(ws, min_width=15, max_width=50)
    
    def _create_vesting_events_sheet(self, wb: Workbook, vesting_events: List[VestingEvent]):
        """Create vesting events sheet."""
        ws = wb.create_sheet("Vesting Events")
        
        # Convert to DataFrame for easier manipulation
        vesting_data = []
        for event in vesting_events:
            vesting_data.append({
                'Vesting Date': event.vest_date.strftime('%d/%m/%Y'),
                'Grant Number': event.grant_number,
                'Shares Vested': event.vested_quantity,
                'FMV per Share (USD)': event.vest_fmv_usd,
                'Exchange Rate': event.exchange_rate,
                'Vesting Value (USD)': event.taxable_gain_usd,
                'Vesting Value (INR)': event.taxable_gain_inr,
                'Financial Year': event.financial_year
            })
        
        df = pd.DataFrame(vesting_data)
        
        # Add title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = "RSU Vesting Events"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add DataFrame to sheet
        data_start_row = 3
        current_row = data_start_row
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=data_start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == data_start_row:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                else:
                    # Apply number formatting based on column
                    if c_idx == 3:  # Shares Vested
                        cell.number_format = '#,##0'
                    elif c_idx == 4:  # FMV per Share (USD)
                        cell.number_format = '"$"#,##0.00'
                    elif c_idx == 5:  # Exchange Rate
                        cell.number_format = '"₹"#,##0.0000'
                    elif c_idx == 6:  # Vesting Value (USD)
                        cell.number_format = '"$"#,##0.00'
                    elif c_idx == 7:  # Vesting Value (INR)
                        cell.number_format = '"₹"#,##0.00'
                
                cell.border = self.border
            current_row = r_idx
        
        # Add total row for USD and INR columns
        if len(vesting_events) > 0:
            total_row = current_row + 1
            
            # Add TOTAL label in first column
            total_cell = ws.cell(row=total_row, column=1, value="TOTAL")
            total_cell.font = Font(bold=True)
            total_cell.border = self.border
            
            # Add dashes for columns 2-5 (Grant Number through Exchange Rate)
            for col in range(2, 6):
                cell = ws.cell(row=total_row, column=col, value="-")
                cell.font = Font(bold=True)
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center")
            
            # Calculate and add total USD amount (column 6)
            total_usd = sum(event.taxable_gain_usd for event in vesting_events)
            usd_total_cell = ws.cell(row=total_row, column=6, value=total_usd)
            usd_total_cell.font = Font(bold=True)
            usd_total_cell.number_format = '"$"#,##0.00'
            usd_total_cell.border = self.border
            
            # Calculate and add total INR amount (column 7)
            total_inr = sum(event.taxable_gain_inr for event in vesting_events)
            inr_total_cell = ws.cell(row=total_row, column=7, value=total_inr)
            inr_total_cell.font = Font(bold=True)
            inr_total_cell.number_format = '"₹"#,##0.00'
            inr_total_cell.border = self.border
            
            # Add dash for Financial Year column (column 8)
            cell = ws.cell(row=total_row, column=8, value="-")
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths with improved calculation
        self._auto_adjust_column_widths(ws, min_width=18, max_width=55)
    
    def _create_sale_events_sheet(self, wb: Workbook, sale_events: List[SaleEvent]):
        """Create sale events sheet."""
        ws = wb.create_sheet("Sale Events")
        
        # Convert to DataFrame
        sale_data = []
        for event in sale_events:
            sale_data.append({
                'Sale Date': event.sale_date.strftime('%d/%m/%Y'),
                'Vest Date': event.acquisition_date.strftime('%d/%m/%Y'),
                'Grant Number': event.grant_number,
                'Shares Sold': event.quantity_sold,
                'Sale Price (USD)': event.sale_price_usd,
                'Rule 115 Exchange Rate': event.capital_gains_exchange_rate or event.exchange_rate_sale,
                'Sale Proceeds (USD)': event.sale_proceeds_usd,
                'Sale Proceeds (INR)': event.sale_proceeds_inr,
                'Cost Basis (USD)': event.cost_basis_usd,
                'Cost Basis (INR)': event.cost_basis_inr,
                'Net Capital Gain (USD)': event.capital_gain_usd,
                'Net Capital Gain (INR)': event.capital_gain_inr,
                'Holding Period': f"{event.holding_period_days} days",
                'Gain Type': event.gain_type,
                'Financial Year': event.financial_year,
                'Sale-Date Exchange Rate': event.exchange_rate_sale,
                'Gross Capital Gain (USD)': event.gross_capital_gain_usd,
                'Gross Capital Gain (INR)': event.gross_capital_gain_inr,
                'Deductible Sale Expense (USD)': event.sale_expense_usd,
                'Deductible Sale Expense (INR)': event.sale_expense_inr,
            })
        
        df = pd.DataFrame(sale_data)
        
        # Add title
        ws.merge_cells('A1:T1')
        title_cell = ws['A1']
        title_cell.value = "RSU Sale Events"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add DataFrame to sheet
        data_start_row = 3
        current_row = data_start_row
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=data_start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == data_start_row:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                else:
                    # Apply number formatting based on column
                    if c_idx == 4:  # Shares Sold
                        cell.number_format = '#,##0'
                    elif c_idx == 5:  # Sale Price (USD)
                        cell.number_format = '"$"#,##0.00'
                    elif c_idx == 6:  # Exchange Rate
                        cell.number_format = '"₹"#,##0.0000'
                    elif c_idx == 7:  # Sale Proceeds (USD)
                        cell.number_format = '"$"#,##0.00'
                    elif c_idx == 8:  # Sale Proceeds (INR)
                        cell.number_format = '"₹"#,##0.00'
                    elif c_idx == 9:  # Cost Basis (USD)
                        cell.number_format = '"$"#,##0.00'
                    elif c_idx == 10:  # Cost Basis (INR)
                        cell.number_format = '"₹"#,##0.00'
                    elif c_idx == 11:  # Capital Gain (USD)
                        cell.number_format = '"$"#,##0.00'
                    elif c_idx == 12:  # Capital Gain (INR)
                        cell.number_format = '"₹"#,##0.00'
                    elif c_idx in (17, 19):
                        cell.number_format = '"$"#,##0.00'
                    elif c_idx in (18, 20):
                        cell.number_format = '"₹"#,##0.00'
                
                cell.border = self.border
            current_row = r_idx
        
        # Add total row for USD and INR columns
        if len(sale_events) > 0:
            total_row = current_row + 1
            
            # Add TOTAL label in first column
            total_cell = ws.cell(row=total_row, column=1, value="TOTAL")
            total_cell.font = Font(bold=True)
            total_cell.border = self.border
            
            # Add dashes for columns 2-6 (Vest Date through Exchange Rate)
            for col in range(2, 7):
                cell = ws.cell(row=total_row, column=col, value="-")
                cell.font = Font(bold=True)
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center")
            
            # Calculate and add totals for relevant columns
            # Sale Proceeds (USD) - column 7
            total_proceeds_usd = sum(event.sale_proceeds_usd for event in sale_events)
            proceeds_usd_cell = ws.cell(row=total_row, column=7, value=total_proceeds_usd)
            proceeds_usd_cell.font = Font(bold=True)
            proceeds_usd_cell.number_format = '"$"#,##0.00'
            proceeds_usd_cell.border = self.border
            
            # Sale Proceeds (INR) - column 8
            total_proceeds_inr = sum(event.sale_proceeds_inr for event in sale_events)
            proceeds_inr_cell = ws.cell(row=total_row, column=8, value=total_proceeds_inr)
            proceeds_inr_cell.font = Font(bold=True)
            proceeds_inr_cell.number_format = '"₹"#,##0.00'
            proceeds_inr_cell.border = self.border
            
            # Cost Basis (USD) - column 9
            total_cost_usd = sum(event.cost_basis_usd for event in sale_events)
            cost_usd_cell = ws.cell(row=total_row, column=9, value=total_cost_usd)
            cost_usd_cell.font = Font(bold=True)
            cost_usd_cell.number_format = '"$"#,##0.00'
            cost_usd_cell.border = self.border
            
            # Cost Basis (INR) - column 10
            total_cost_inr = sum(event.cost_basis_inr for event in sale_events)
            cost_inr_cell = ws.cell(row=total_row, column=10, value=total_cost_inr)
            cost_inr_cell.font = Font(bold=True)
            cost_inr_cell.number_format = '"₹"#,##0.00'
            cost_inr_cell.border = self.border
            
            # Capital Gain (USD) - column 11
            total_gain_usd = sum(event.capital_gain_usd for event in sale_events)
            gain_usd_cell = ws.cell(row=total_row, column=11, value=total_gain_usd)
            gain_usd_cell.font = Font(bold=True)
            gain_usd_cell.number_format = '"$"#,##0.00'
            gain_usd_cell.border = self.border
            
            # Capital Gain (INR) - column 12
            total_gain_inr = sum(event.capital_gain_inr for event in sale_events)
            gain_inr_cell = ws.cell(row=total_row, column=12, value=total_gain_inr)
            gain_inr_cell.font = Font(bold=True)
            gain_inr_cell.number_format = '"₹"#,##0.00'
            gain_inr_cell.border = self.border
            
            # Add dashes for non-total columns (holding/type/FY/sale-date rate).
            for col in range(13, 17):
                cell = ws.cell(row=total_row, column=col, value="-")
                cell.font = Font(bold=True)
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center")

            extra_totals = {
                17: sum(event.gross_capital_gain_usd for event in sale_events),
                18: sum(event.gross_capital_gain_inr for event in sale_events),
                19: sum(event.sale_expense_usd for event in sale_events),
                20: sum(event.sale_expense_inr for event in sale_events),
            }
            for col, value in extra_totals.items():
                cell = ws.cell(row=total_row, column=col, value=value)
                cell.font = Font(bold=True)
                cell.number_format = '"$"#,##0.00' if col in (17, 19) else '"₹"#,##0.00'
                cell.border = self.border
        
        # Auto-adjust column widths with improved calculation
        self._auto_adjust_column_widths(ws, min_width=18, max_width=50)
    
    def _create_bank_reconciliation_sheet(self, wb: Workbook, sale_events: List[SaleEvent], bank_transactions: List[Dict]):
        """Create bank reconciliation sheet."""
        ws = wb.create_sheet("Bank Reconciliation")
        
        # Group sale events by sale date for reconciliation
        from collections import defaultdict
        
        sale_by_date = defaultdict(list)
        for event in sale_events:
            sale_by_date[event.sale_date].append(event)
        
        # Create reconciliation data
        recon_data = []
        for sale_date, events in sale_by_date.items():
            total_usd = sum(event.sale_proceeds_usd for event in events)
            
            # Find matching bank transaction
            bank_match = next(
                (
                    tx for tx in bank_transactions
                    if tx.get('sale_date') == sale_date
                ),
                None,
            )
            
            expected_inr = total_usd * events[0].exchange_rate_sale
            bank_received_inr = bank_match.get('actual_received', 0) if bank_match else 0
            # Net Difference: Final Received - Expected (positive=gain, negative=loss)
            net_difference = bank_received_inr - expected_inr if bank_match else 0
            
            recon_data.append({
                'Sale Date': sale_date.strftime('%d/%m/%Y'),
                'Expected USD': f"${total_usd:.2f}",
                'Expected INR': f"₹{expected_inr:.2f}",
                'Bank Received USD': f"${bank_match.get('bank_usd_amount', 0):.2f}" if bank_match else "Not Found",
                'Bank Received INR': f"₹{bank_received_inr:.2f}" if bank_match else "Not Found",
                'Net Difference INR': f"₹{net_difference:.2f}" if bank_match else "N/A",
                'Deductible Sale Expense': f"${bank_match.get('sale_expense_usd', 0):.2f}" if bank_match else "N/A",
                'Exchange Rate Diff': f"₹{bank_match.get('exchange_rate_gain_loss', 0):.2f}" if bank_match else "N/A"
            })
        
        df = pd.DataFrame(recon_data)
        
        # Add title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = "Bank Reconciliation"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add DataFrame to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == 3:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                
                cell.border = self.border
        
        # Auto-adjust column widths with improved calculation
        self._auto_adjust_column_widths(ws, min_width=18, max_width=50)
    
    def _create_fa_summary_sheet(self, wb: Workbook, summary: FADeclarationSummary, calendar_year: str):
        """Create FA summary sheet."""
        ws = wb.create_sheet("FA Summary")
        
        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f"Foreign Assets Declaration - {calendar_year or 'All Years'}"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Summary data - store raw values for proper number formatting
        summary_data = [
            ["", "Metric", "Value", "Notes"],
            ["Holdings", "Total Vested Shares", f"{summary.total_vested_shares:.0f} shares", ""],
            ["", "Closing Balance", summary.closing_balance_inr, "As on Dec 31"],
            ["", "Peak Balance", summary.peak_balance_inr, f"Peak Date: {summary.peak_balance_date or 'N/A'}"],
            ["", "Opening Balance", summary.opening_balance_inr, "As on Jan 1"],
            ["", "", "", ""],
            ["Exchange Rates", "Year-end Rate", summary.year_end_exchange_rate, "Dec 31 rate"],
            ["", "Opening Rate", summary.opening_exchange_rate, "Jan 1 rate"],
            ["", "Peak Rate", summary.peak_exchange_rate, "Highest during year"],
            ["", "", "", ""],
            ["Declaration", "Declaration Required?", "YES" if summary.closing_balance_inr >= summary.fa_declaration_threshold_inr else "NO", f"Threshold: ₹{summary.fa_declaration_threshold_inr:,.0f}"],
            ["", "Total Value (USD)", summary.vested_holdings_usd, "At year-end rates"],
            ["", "Total Value (INR)", summary.vested_holdings_inr, "At year-end rates"]
        ]
        
        # Add data to sheet
        for row_idx, row_data in enumerate(summary_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply number formatting for currency values
                if col_idx == 3 and isinstance(value, (int, float)) and value != 0:  # Value column
                    metric_name = str(row_data[1]) if len(row_data) > 1 else ""
                    if "Balance" in metric_name or "Value (INR)" in metric_name:
                        cell.number_format = '"₹"#,##0.00'
                    elif "Value (USD)" in metric_name:
                        cell.number_format = '"$"#,##0.00'
                    elif "Rate" in metric_name and "Exchange" in str(row_data[0]):
                        cell.number_format = '"₹"#,##0.0000"/USD"'
                
                # Style header row
                if row_idx == 3:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                
                # Style declaration row
                elif "Declaration Required" in str(value):
                    cell.font = Font(bold=True, size=12)
                    if "YES" in str(row_data[2]):
                        cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
                
                cell.border = self.border
        
        # Auto-adjust column widths with improved calculation
        self._auto_adjust_column_widths(ws, min_width=15, max_width=50)
    
    def _create_equity_holdings_sheet(self, wb: Workbook, equity_holdings: List[EquityHolding]):
        """Create equity holdings sheet."""
        ws = wb.create_sheet("Equity Holdings")
        
        # Convert to DataFrame - store raw values for proper number formatting
        holdings_data = []
        for holding in equity_holdings:
            holdings_data.append({
                'Grant Number': holding.grant_number or 'N/A',
                'Vest Date': holding.vest_date.strftime('%d/%m/%Y') if holding.vest_date else 'N/A',
                'Holding Date': holding.holding_date.strftime('%d/%m/%Y'),
                'Shares Held': holding.quantity,
                'Cost Basis per Share (USD)': holding.cost_basis_usd_per_share,
                'Market Price per Share (USD)': holding.market_value_usd_per_share,
                'Total Cost Basis (USD)': holding.cost_basis_usd_total,
                'Total Market Value (USD)': holding.market_value_usd_total,
                'Total Market Value (INR)': holding.market_value_inr_total,
                'Exchange Rate': holding.exchange_rate,
                'Unrealized Gain (USD)': holding.unrealized_gain_usd,
                'Unrealized Gain (INR)': holding.unrealized_gain_inr,
                'Calendar Year': holding.calendar_year
            })
        
        df = pd.DataFrame(holdings_data)
        
        # Add title
        ws.merge_cells('A1:M1')
        title_cell = ws['A1']
        title_cell.value = "Current Equity Holdings"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add DataFrame to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == 3:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                else:
                    # Apply number formatting based on column
                    if c_idx == 4:  # Shares Held
                        cell.number_format = '#,##0'
                    elif c_idx == 5:  # Cost Basis per Share (USD)
                        cell.number_format = '"$"#,##0.00'
                    elif c_idx == 6:  # Market Price per Share (USD)
                        cell.number_format = '"$"#,##0.00'
                    elif c_idx == 7:  # Total Cost Basis (USD)
                        cell.number_format = '"$"#,##0.00'
                    elif c_idx == 8:  # Total Market Value (USD)
                        cell.number_format = '"$"#,##0.00'
                    elif c_idx == 9:  # Total Market Value (INR)
                        cell.number_format = '"₹"#,##0.00'
                    elif c_idx == 10:  # Exchange Rate
                        cell.number_format = '"₹"#,##0.0000'
                    elif c_idx == 11:  # Unrealized Gain (USD)
                        cell.number_format = '"$"#,##0.00'
                    elif c_idx == 12:  # Unrealized Gain (INR)
                        cell.number_format = '"₹"#,##0.00'
                
                cell.border = self.border
        
        # Add total rows for USD and INR currency columns
        if equity_holdings:
            total_row = len(df) + 4  # +3 for title and header, +1 for data
            
            # Add TOTAL label in first column
            total_cell = ws.cell(row=total_row, column=1, value="TOTAL")
            total_cell.font = Font(bold=True)
            total_cell.border = self.border
            
            # Add dashes for non-numeric columns
            for col in [2, 3, 13]:  # Vest Date, Holding Date, Calendar Year
                cell = ws.cell(row=total_row, column=col, value="-")
                cell.font = Font(bold=True)
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center")
            
            # Calculate and add totals for currency columns
            total_shares = sum(h.quantity for h in equity_holdings)
            total_cost_basis_usd = sum(h.cost_basis_usd_total for h in equity_holdings)
            total_market_value_usd = sum(h.market_value_usd_total for h in equity_holdings)
            total_market_value_inr = sum(h.market_value_inr_total for h in equity_holdings)
            total_unrealized_gain_usd = sum(h.unrealized_gain_usd for h in equity_holdings)
            total_unrealized_gain_inr = sum(h.unrealized_gain_inr for h in equity_holdings)
            
            # Column 4: Shares Held
            shares_cell = ws.cell(row=total_row, column=4, value=total_shares)
            shares_cell.font = Font(bold=True)
            shares_cell.number_format = '#,##0'
            shares_cell.border = self.border
            shares_cell.alignment = Alignment(horizontal="center")
            
            # Skip per-share values (columns 5-6) with dashes
            for col in [5, 6]:
                cell = ws.cell(row=total_row, column=col, value="-")
                cell.font = Font(bold=True)
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center")
            
            # Column 7: Total Cost Basis (USD)
            cost_cell = ws.cell(row=total_row, column=7, value=total_cost_basis_usd)
            cost_cell.font = Font(bold=True)
            cost_cell.number_format = '"$"#,##0.00'
            cost_cell.border = self.border
            
            # Column 8: Total Market Value (USD)
            market_usd_cell = ws.cell(row=total_row, column=8, value=total_market_value_usd)
            market_usd_cell.font = Font(bold=True)
            market_usd_cell.number_format = '"$"#,##0.00'
            market_usd_cell.border = self.border
            
            # Column 9: Total Market Value (INR)
            market_inr_cell = ws.cell(row=total_row, column=9, value=total_market_value_inr)
            market_inr_cell.font = Font(bold=True)
            market_inr_cell.number_format = '"₹"#,##0.00'
            market_inr_cell.border = self.border
            
            # Column 10: Exchange Rate (dash)
            rate_cell = ws.cell(row=total_row, column=10, value="-")
            rate_cell.font = Font(bold=True)
            rate_cell.border = self.border
            rate_cell.alignment = Alignment(horizontal="center")
            
            # Column 11: Unrealized Gain (USD)
            gain_usd_cell = ws.cell(row=total_row, column=11, value=total_unrealized_gain_usd)
            gain_usd_cell.font = Font(bold=True)
            gain_usd_cell.number_format = '"$"#,##0.00'
            gain_usd_cell.border = self.border
            
            # Column 12: Unrealized Gain (INR)
            gain_inr_cell = ws.cell(row=total_row, column=12, value=total_unrealized_gain_inr)
            gain_inr_cell.font = Font(bold=True)
            gain_inr_cell.number_format = '"₹"#,##0.00'
            gain_inr_cell.border = self.border
        
        # Auto-adjust column widths with improved calculation
        self._auto_adjust_column_widths(ws, min_width=18, max_width=50)
    
    def _create_vest_wise_details_sheet(self, wb: Workbook, vest_wise_details: List[VestWiseDetails]):
        """Create vest-wise details sheet."""
        ws = wb.create_sheet("Vest-wise Details")
        
        # Convert to DataFrame - store raw values for proper number formatting
        vest_data = []
        for detail in vest_wise_details:
            vest_data.append({
                'Grant Number': detail.grant_number,
                'Vest Date': detail.vest_date.strftime('%d/%m/%Y'),
                'Initial Value (INR)': detail.initial_value_inr,
                'Peak Value (INR)': detail.peak_value_inr,
                'Closing Value (INR)': detail.closing_value_inr,
                'Gross Income (INR)': detail.gross_income_received,
                'Sale Proceeds (INR)': detail.gross_proceeds_inr,
                'Shares at Year-end': detail.closing_shares,
                'Shares Sold': detail.shares_sold
            })
        
        df = pd.DataFrame(vest_data)
        
        # Add title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = "Vest-wise Investment Details"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add DataFrame to sheet
        data_start_row = 3
        current_row = data_start_row
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=data_start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == data_start_row:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                else:
                    # Apply number formatting based on column
                    if c_idx in [3, 4, 5, 6, 7]:  # All INR currency columns
                        cell.number_format = '"₹"#,##0.00'
                    elif c_idx in [8, 9]:  # Shares columns
                        cell.number_format = '#,##0'
                
                cell.border = self.border
            current_row = r_idx
        
        # Add total rows for currency columns
        if vest_wise_details:
            total_row = current_row + 1
            
            # Column mapping based on DataFrame structure:
            # 1: Grant Number, 2: Vest Date, 3: Initial Value, 4: Peak Value, 
            # 5: Closing Value, 6: Gross Income, 7: Sale Proceeds, 8: Shares at Year-end, 9: Shares Sold
            
            # Calculate totals for currency columns
            total_initial_value = sum(detail.initial_value_inr for detail in vest_wise_details)
            total_peak_value = sum(detail.peak_value_inr for detail in vest_wise_details)
            total_closing_value = sum(detail.closing_value_inr for detail in vest_wise_details)
            total_gross_income = sum(detail.gross_income_received for detail in vest_wise_details)
            total_sale_proceeds = sum(detail.gross_proceeds_inr for detail in vest_wise_details if detail.gross_proceeds_inr > 0)
            total_shares_yearend = sum(detail.closing_shares for detail in vest_wise_details)
            total_shares_sold = sum(detail.shares_sold for detail in vest_wise_details)
            
            # Add TOTAL label in first column
            total_cell = ws.cell(row=total_row, column=1, value="TOTAL")
            total_cell.font = Font(bold=True)
            total_cell.border = self.border
            
            # Add dash for Vest Date column
            dash_cell = ws.cell(row=total_row, column=2, value="-")
            dash_cell.font = Font(bold=True)
            dash_cell.border = self.border
            dash_cell.alignment = Alignment(horizontal="center")
            
            # Column 3: Initial Value (INR)
            initial_cell = ws.cell(row=total_row, column=3, value=total_initial_value)
            initial_cell.font = Font(bold=True)
            initial_cell.number_format = '"₹"#,##0.00'
            initial_cell.border = self.border
            
            # Column 4: Peak Value (INR)
            peak_cell = ws.cell(row=total_row, column=4, value=total_peak_value)
            peak_cell.font = Font(bold=True)
            peak_cell.number_format = '"₹"#,##0.00'
            peak_cell.border = self.border
            
            # Column 5: Closing Value (INR)
            closing_cell = ws.cell(row=total_row, column=5, value=total_closing_value)
            closing_cell.font = Font(bold=True)
            closing_cell.number_format = '"₹"#,##0.00'
            closing_cell.border = self.border
            
            # Column 6: Gross Income (INR)
            income_cell = ws.cell(row=total_row, column=6, value=total_gross_income)
            income_cell.font = Font(bold=True)
            income_cell.number_format = '"₹"#,##0.00'
            income_cell.border = self.border
            
            # Column 7: Sale Proceeds (INR)
            proceeds_cell = ws.cell(row=total_row, column=7, value=total_sale_proceeds)
            proceeds_cell.font = Font(bold=True)
            proceeds_cell.number_format = '"₹"#,##0.00'
            proceeds_cell.border = self.border
            
            # Column 8: Shares at Year-end
            yearend_shares_cell = ws.cell(row=total_row, column=8, value=total_shares_yearend)
            yearend_shares_cell.font = Font(bold=True)
            yearend_shares_cell.number_format = '#,##0'
            yearend_shares_cell.border = self.border
            yearend_shares_cell.alignment = Alignment(horizontal="center")
            
            # Column 9: Shares Sold
            sold_shares_cell = ws.cell(row=total_row, column=9, value=total_shares_sold)
            sold_shares_cell.font = Font(bold=True)
            sold_shares_cell.number_format = '#,##0'
            sold_shares_cell.border = self.border
            sold_shares_cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths with improved calculation
        self._auto_adjust_column_widths(ws, min_width=18, max_width=50)
    
    def _create_company_details_sheet(self, wb: Workbook):
        """Create company and depository account details sheet for FA filing."""
        from ..data.models import create_default_company_records
        
        ws = wb.create_sheet("Company & Account Details")
        
        # Get company and depository account details
        employer_company, foreign_company, depository_account = create_default_company_records()
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "Company & Depository Account Details for FA Filing"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Company Information Section
        row = 3
        ws.merge_cells(f'A{row}:F{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "📋 Company Information"
        section_cell.font = Font(bold=True, size=14)
        section_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        section_cell.alignment = Alignment(horizontal="center")
        
        # Company headers
        row += 1
        headers = ["Type", "Company Name", "Address", "City/State", "ZIP/PIN", "ID/TAN"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = self.border
        
        # Employer company data
        row += 1
        employer_data = [
            "Employer (India)",
            employer_company.company_name,
            employer_company.address_line1,
            f"{employer_company.city}, {employer_company.state}",
            employer_company.pin_code,
            employer_company.tan
        ]
        for col, value in enumerate(employer_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.border
        
        # Foreign company data
        row += 1
        foreign_data = [
            "Foreign Entity",
            foreign_company.company_name,
            foreign_company.address_line1,
            f"{foreign_company.city}, {foreign_company.state_province}",
            foreign_company.zip_code,
            f"Country Code: {foreign_company.country_code}"
        ]
        for col, value in enumerate(foreign_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.border
        
        # Depository Account Information Section
        row += 3
        ws.merge_cells(f'A{row}:F{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "🏛️ Foreign Depository Account Information"
        section_cell.font = Font(bold=True, size=14)
        section_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        section_cell.alignment = Alignment(horizontal="center")
        
        # Account headers
        row += 1
        account_headers = ["Institution", "Address", "Account Number", "Status", "Opening Date", "Country"]
        for col, header in enumerate(account_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = self.border
        
        # Depository account data
        row += 1
        institution_address = f"{depository_account.institution_address}, {depository_account.institution_city}, {depository_account.institution_state} {depository_account.institution_zip}"
        account_data = [
            depository_account.institution_name,
            institution_address,
            depository_account.account_number,
            depository_account.account_status,
            depository_account.account_opening_date.strftime("%d/%m/%Y"),
            f"{depository_account.institution_country} (Code: {depository_account.institution_country_code})"
        ]
        for col, value in enumerate(account_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.border
        
        # ITR Schedule References Section
        row += 3
        ws.merge_cells(f'A{row}:F{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "📄 ITR Schedule References"
        section_cell.font = Font(bold=True, size=14)
        section_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        section_cell.alignment = Alignment(horizontal="center")
        
        # ITR reference notes
        row += 2
        itr_notes = [
            "Schedule A1 - Foreign Depository Accounts:",
            f"  • Institution: {depository_account.institution_name}",
            f"  • Account: {depository_account.account_number} ({depository_account.account_status})",
            f"  • Address: {institution_address}",
            "",
            "Schedule A3 - Foreign Equity and Debt Interest:",
            f"  • Entity: {foreign_company.company_name}",
            f"  • Address: {foreign_company.address_line1}, {foreign_company.city}, {foreign_company.state_province} {foreign_company.zip_code}",
            f"  • Country: {foreign_company.country_name} (Code: {foreign_company.country_code})",
            f"  • Nature: {foreign_company.nature_of_entity}",
            "",
            "💡 Important Notes:",
            "  • These details are extracted from your previous ITR filing",
            "  • Country Code 2 = United States of America",
            "  • Use these exact details when filling FA schedules to maintain consistency"
        ]
        
        for note in itr_notes:
            ws.cell(row=row, column=1, value=note)
            row += 1
        
        # Auto-adjust column widths with improved calculation
        self._auto_adjust_column_widths(ws, min_width=18, max_width=55)
    
    def _auto_adjust_column_widths(self, ws, min_width: int = 12, max_width: int = 60):
        """Simple but robust column width adjustment to prevent ######### display."""
        from openpyxl.utils import get_column_letter
        
        # Iterate through columns by index to avoid merged cell issues
        for col_idx, column_cells in enumerate(ws.columns, 1):
            # Get column letter from column index
            column_letter = get_column_letter(col_idx)
            
            # Get column header to determine content type, skip merged cells
            header_text = ""
            for cell in column_cells:
                # Skip merged cells by checking if cell has column_letter attribute
                # Merged cells don't have this attribute
                if not hasattr(cell, 'column_letter'):
                    continue
                if cell.value and isinstance(cell.value, str):
                    header_text = str(cell.value).lower()
                    break
            
            # Set width based on content type with very generous defaults
            if any(word in header_text for word in ['date', 'period', 'year']):
                # Date columns - need space for "31/01/2025" format
                width = 15
            elif any(word in header_text for word in ['number', 'grant']):
                # Grant/ID numbers - medium width  
                width = 18
            elif any(word in header_text for word in ['shares', 'quantity']):
                # Share quantities - medium width for numbers like "1,234"
                width = 12
            elif any(word in header_text for word in ['price', 'rate']) and 'exchange' not in header_text:
                # Stock prices - need space for "$563.54" format
                width = 15
            elif 'exchange' in header_text and 'rate' in header_text:
                # Exchange rates - need space for "₹86.6414" format  
                width = 18
            elif any(word in header_text for word in ['proceeds', 'basis', 'gain', 'value', 'amount', 'income']):
                # Currency amounts - need lots of space for "₹1,099,543.25" format
                if 'inr' in header_text or '₹' in header_text:
                    width = 25  # INR amounts tend to be larger
                else:
                    width = 20  # USD amounts
            elif any(word in header_text for word in ['type', 'term']):
                # Text classifications like "Short-term"
                width = 15
            else:
                # Default generous width for unknown columns
                width = 18
            
            # Apply the calculated width
            ws.column_dimensions[column_letter].width = width
